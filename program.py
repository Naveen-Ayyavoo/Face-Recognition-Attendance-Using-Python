import face_recognition
import cv2
import numpy as np
import os
from datetime import datetime, time
import threading
import schedule
import time as tm
import openpyxl
from openpyxl.styles import Alignment


# Class schedule definition
class_schedule = {
    "Class 1": {"start": time(9, 0), "end": time(9, 50)},
    "Class 2": {"start": time(9, 50), "end": time(10, 40)},
    "Class 3": {"start": time(10, 55), "end": time(11, 45)},
    "Class 4": {"start": time(11, 45), "end": time(12, 30)},
    "Class 5": {"start": time(13, 15), "end": time(14, 00)},
    "Class 6": {"start": time(14, 00), "end": time(14, 45)},
    "Class 7": {"start": time(14, 45), "end": time(15, 30)}
}

# Global variables
exit_flag = False
current_session = None
attendance_system_running = False
video_capture = None
process_thread = None
display_thread = None

# Function to get daily CSV filename
def get_daily_xlsx_filename():
    return f"IT_Attendance_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

# Function to check if CSV exists and create with headers if not
def initialize_xlsx():
    xlsx_filename = get_daily_xlsx_filename()
    if not os.path.isfile(xlsx_filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        headers = ["Register Number", "Course", "Name", "Date"]
        for session in class_schedule.keys():
            headers.append(session)
        ws.append(headers)
        wb.save(xlsx_filename)
        print(f"✅ Created new attendance file: {xlsx_filename}")
    return xlsx_filename

def read_current_attendance():
    xlsx_filename = get_daily_xlsx_filename()
    attendance_data = {}

    if not os.path.isfile(xlsx_filename):
        return attendance_data

    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return attendance_data

    headers = rows[0]
    class_indices = {}
    for session in class_schedule.keys():
        if session in headers:
            class_indices[session] = headers.index(session)

    for row in rows[1:]:
        if len(row) < len(headers):
            continue
        name = row[2]
        attendance_data[name] = {
            "reg_number": row[0],
            "course": row[1],
            "date": row[3],
            "classes": {}
        }
        for session, idx in class_indices.items():
            if idx < len(row):
                attendance_data[name]["classes"][session] = row[idx]
            else:
                attendance_data[name]["classes"][session] = "ABSENT"
        attendance_data[name]["total"] = sum(1 for status in attendance_data[name]["classes"].values() if status == "PRESENT")
    return attendance_data

def update_attendance(name, reg_number, course, session_name, status="PRESENT"):
    current_date = datetime.now().strftime("%d-%m-%Y")
    attendance_data = read_current_attendance()
    if name in attendance_data:
        attendance_data[name]["classes"][session_name] = status
        attendance_data[name]["total"] = sum(1 for s in attendance_data[name]["classes"].values() if s == "PRESENT")
    else:
        attendance_data[name] = {
            "reg_number": reg_number,
            "course": course,
            "date": current_date,
            "classes": {session: "ABSENT" for session in class_schedule.keys()},
            "total": 0
        }
        attendance_data[name]["classes"][session_name] = status
        attendance_data[name]["total"] = sum(1 for s in attendance_data[name]["classes"].values() if s == "PRESENT")
    write_attendance_data(attendance_data)
    return attendance_data

def write_attendance_data(attendance_data):
    xlsx_filename = get_daily_xlsx_filename()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["Register Number", "Course", "Name", "Date"]
    for session in class_schedule.keys():
        headers.append(session)
    headers.append("Total Present")
    ws.append(headers)

    current_row = 2
    for name, data in attendance_data.items():
        if not name or name == "None":
            continue
        row = [
            data.get("reg_number", "Unknown"),
            data.get("course", "Unknown"),
            name,
            data.get("date", datetime.now().strftime("%d-%m-%Y"))
        ]
        for session in class_schedule.keys():
            row.append(data["classes"].get(session, "ABSENT"))
        row.append(data.get("total", 0))
        ws.append(row)
        current_row += 1

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value else 0
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Center align all cells
    alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment

    wb.save(xlsx_filename)
    print(f"✅ Attendance Excel file saved: {xlsx_filename}")

# Load known faces
def load_known_faces():
    photos_dir = "student_faces"
    known_face_encodings = []
    known_faces_names = []
    known_reg_numbers = []
    known_courses = []

    if not os.path.exists(photos_dir):
        os.makedirs(photos_dir)
        print(f"Created directory: {photos_dir}")
        print("Please add student folders and photos before running attendance")
        return [], [], [], []

    for folder_name in os.listdir(photos_dir):
        folder_path = os.path.join(photos_dir, folder_name)
        if os.path.isdir(folder_path):
            try:
                reg_number, course, name = folder_name.split('_', 2)
            except ValueError:
                print(f"❌ Folder name {folder_name} is not properly formatted. Expected format: RegNo_Course_Name")
                continue

            loaded_any = False  # To check if at least one face was loaded

            for filename in os.listdir(folder_path):
                if filename.endswith(('.jpeg', '.jpg', '.png')):
                    image_path = os.path.join(folder_path, filename)
                    image = face_recognition.load_image_file(image_path)
                    try:
                        face_encoding = face_recognition.face_encodings(image)[0]
                        known_face_encodings.append(face_encoding)
                        known_faces_names.append(name)
                        known_reg_numbers.append(reg_number)
                        known_courses.append(course)
                        loaded_any = True
                    except IndexError:
                        print(f"❌ Warning: No face found in {filename}, skipping.")

            if loaded_any:
                print(f"✅ Loaded faces for {folder_name}")

    return known_face_encodings, known_faces_names, known_reg_numbers, known_courses



# Start the attendance system
def start_attendance_system(session_name):
    global exit_flag, attendance_system_running, video_capture, process_thread, display_thread, current_session
    
    if attendance_system_running:
        print(f"⚠️ Attendance system already running for {current_session}")
        return
        
    print(f"🔄 Starting attendance system for {session_name}")
    current_session = session_name
    exit_flag = False
    attendance_system_running = True
    
    # Initialize camera
    video_capture = cv2.VideoCapture(0)
    
    # Load known faces
    known_face_encodings, known_faces_names, known_reg_numbers, known_courses = load_known_faces()
    students = [{"name": n, "reg_number": r, "course": c} for n, r, c in zip(known_faces_names, known_reg_numbers, known_courses)]
    
    # Initialize CSV file
    csv_filename = initialize_xlsx()
    
    # Initialize attendance data
    attendance_data = read_current_attendance()
    
    # Make sure all students are in the attendance data
    for student in students:
        name = student["name"]
        if name not in attendance_data:
            current_date = datetime.now().strftime("%d-%m-%Y")
            attendance_data[name] = {
                "reg_number": student["reg_number"],
                "course": student["course"],
                "date": current_date,
                "classes": {session: "ABSENT" for session in class_schedule.keys()},
                "total": 0
            }
    
    # Update attendance data to ensure all students are marked as absent for this session if not already marked
    for name, data in attendance_data.items():
        if session_name not in data["classes"]:
            data["classes"][session_name] = "ABSENT"
    
    # Write initial attendance data
    write_attendance_data(attendance_data)
    
    face_locations = []
    face_encodings = []
    face_names = []
    face_confidences = []
    face_courses = []
    recognized_students = set()

    CONFIDENCE_THRESHOLD = 0.60  # 60% confidence

    def process_frame():
        nonlocal face_names, face_locations, face_encodings, face_confidences, face_courses, recognized_students

        while not exit_flag:
            ret, frame = video_capture.read()
            if not ret or exit_flag:
                break

            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            face_confidences = []
            face_courses = []
            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                name, reg_number, course, confidence_display = "Unknown", "N/A", "N/A", "N/A"

                if len(face_distances) > 0:
                    best_match_index = np.argmin(face_distances)
                    confidence = (1 - face_distances[best_match_index]) * 100  # Convert to percentage

                    if matches[best_match_index] and confidence >= (CONFIDENCE_THRESHOLD * 100):
                        name = known_faces_names[best_match_index]
                        reg_number = known_reg_numbers[best_match_index]
                        course = known_courses[best_match_index]
                        confidence_display = f"{confidence:.1f}%"

                        # Mark attendance if not already recognized in this session
                        if name not in recognized_students:
                            recognized_students.add(name)
                            
                            # Update attendance record
                            nonlocal attendance_data
                            attendance_data = update_attendance(name, reg_number, course, session_name)
                            print(f"✅ Marked attendance for {name} (Confidence: {confidence_display}) in {session_name}")

                face_names.append(name)
                face_confidences.append(confidence_display)
                face_courses.append(course)
            
            tm.sleep(0.1)  # Prevent high CPU usage

    def display_frame(session_start_time):
        while not exit_flag:
            ret, frame = video_capture.read()
            if not ret or exit_flag:
                break

            # Add session name to frame
            cv2.putText(frame, f"Session: {session_name}", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)
            
            # Calculate and display countdown timer
            elapsed = int(tm.time() - session_start_time)
            remaining = max(0, 600 - elapsed)  # 600 seconds = 10 minutes
            mins, secs = divmod(remaining, 60)
            timer_str = f"Remaining Time: {mins:02d}:{secs:02d}"
            cv2.putText(frame, timer_str, (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 0), 2)
            
            for (top, right, bottom, left), name, confidence, course in zip(face_locations, face_names, face_confidences, face_courses):
                top, right, bottom, left = top * 4, right * 4, bottom * 4, left * 4
                color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
                cv2.rectangle(frame, (left, top), (right, bottom), color, 2)

                text = f"{course} - {name} ({confidence})" if name != "Unknown" else "Unknown"
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
                cv2.putText(frame, text, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 0.5, (0, 0, 0), 1)

            cv2.imshow("Attendance System", frame)
            cv2.waitKey(1)
            tm.sleep(0.03)  # Prevent high CPU usage

    # In start_attendance_system, before starting the threads:
    session_start_time = tm.time()
    process_thread = threading.Thread(target=process_frame, daemon=True)
    display_thread = threading.Thread(target=display_frame, args=(session_start_time,), daemon=True)
    process_thread.start()
    display_thread.start()
    
    # Schedule system to stop after 10 minutes
    tm.sleep(600)  # Run for 10 minutes
    stop_attendance_system(students, session_name)

# Stop the attendance system
def stop_attendance_system(students, session_name):
    global exit_flag, attendance_system_running, video_capture, current_session
    
    if not attendance_system_running:
        return
        
    print(f"🛑 Stopping attendance system for {current_session}")
    exit_flag = True
    attendance_system_running = False
    
    # Wait for threads to finish
    if process_thread:
        process_thread.join(timeout=5)
    if display_thread:
        display_thread.join(timeout=5)
    
    # Read current attendance data
    attendance_data = read_current_attendance()
    
    # Mark absent students
    current_date = datetime.now().strftime("%d-%m-%Y")
    for student in students:
        name = student["name"]
        if name not in attendance_data:
            # Add new student record
            attendance_data[name] = {
                "reg_number": student["reg_number"],
                "course": student["course"],
                "date": current_date,
                "classes": {session: "ABSENT" for session in class_schedule.keys()},
                "total": 0
            }
        
        # Make sure session status is recorded
        if session_name not in attendance_data[name]["classes"]:
            attendance_data[name]["classes"][session_name] = "ABSENT"
    
    # Write final attendance data
    write_attendance_data(attendance_data)
    
    # Print attendance summary
    print(f"\n📊 Attendance Summary for {session_name}:")
    print("-" * 50)
    print(f"{'Name':<20} {'Course':<10} {'Status':<10} {'Total Classes'}")
    print("-" * 50)
    for name, data in attendance_data.items():
        status = data["classes"].get(session_name, "ABSENT")
        print(f"{name:<20} {data['course']:<10} {status:<10} {data['total']}/{len(class_schedule)}")
    print("-" * 50)
    
    # Cleanup
    if video_capture:
        video_capture.release()
    try:
        cv2.destroyAllWindows()
    except KeyboardInterrupt:
        pass
    print(f"💤 Attendance system for {current_session} completed")

# Schedule all sessions
def schedule_all_sessions():
    for session_name, times in class_schedule.items():
        start_time = times["start"].strftime("%H:%M")
        schedule.every().day.at(start_time).do(start_attendance_system, session_name)
        print(f"📅 Scheduled {session_name} to start at {start_time}")

# Main function
def main():
    print("🔄 Starting Automated Attendance System")
    print("📋 Class Schedule:")
    for session, times in class_schedule.items():
        print(f"  • {session}: {times['start'].strftime('%H:%M')} to {times['end'].strftime('%H:%M')}")
    
    schedule_all_sessions()
    
    print("⏰ System will automatically start and stop according to schedule")
    print("Press Ctrl+C to exit the program")
    
    try:
        while True:
            schedule.run_pending()
            tm.sleep(1)
    except KeyboardInterrupt:
        print("👋 Program terminated by user")
        if attendance_system_running:
            exit_flag = True
            if video_capture:
                video_capture.release()
            try:
                cv2.destroyAllWindows()
            except KeyboardInterrupt:
                pass

if __name__ == "__main__":
    main()